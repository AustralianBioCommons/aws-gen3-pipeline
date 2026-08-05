"""Glue job: ingest filled gen3-metadata-templates workbooks into bronze.

The submission path for people who are not data engineers:

    g3mt generate <schema> <node> -o template.xlsx     (a researcher gets a workbook)
    ...fill it in Excel, dropdowns and all...
    g3mt validate template.xlsx --schema <schema>      (optional, catches errors early)
    aws s3 cp template.xlsx s3://<bronze-bucket>/submissions/<study_id>/
                                                      (deposit — that is the whole handover)
    -> this job                                       (one bronze Iceberg table per node sheet)

Everything about the workbook layout is fixed by `g3mt`, so this job does not
guess:

* The **`_g3mt` sheet** carries the provenance block (`g3mt_version`,
  `schema_file`, `target_node`, `path`) followed by an explicit
  **node -> sheet map**. That map is the authoritative list of data sheets —
  `Instructions`, `Dictionary` and `_lists` are simply never in it, so no
  name-based exclusion list can drift out of date.
* On each node sheet, **row 1 is the header, row 2 is the type/required hint
  row, and data starts at row 3.** Row 2 must be dropped or every table gains a
  junk first record reading "string — required".
* `g3mt generate` provisions blank data rows (5000 by default), so the vast
  majority of rows in a fresh workbook are empty and are discarded here.

Bronze stays deliberately dumb (see docs/DATA_LAYERS.md): values land as the
strings the researcher typed. Multi-value cells (`list, separate with ";"`) are
**not** split, enums are not checked, and links are not resolved — that is
silver's job. What this job does add is provenance, so any bronze row can be
traced back to the exact file, sheet and spreadsheet row it came from.

**Re-depositing the same workbook is a no-op.** Each row carries a `row_hash`
over its content plus its source coordinates, and the write is a MERGE on that
hash. This is deliberate: an additive-by-default ingest silently doubles its
corpus the first time someone re-uploads a file, which is exactly the failure
the indexd registry hit in the legacy pipeline (46,598 rows for 23,295 files).

Reads:  s3://<bronze-bucket>/<prefix>/<study_id>/*.xlsx  (prefix defaults to
        `submissions`; study_id is taken from the first path segment under it)
Writes: <project>_<env>_bronze_db.bronze_<study_id>_<node>

Names come from the env's SSM tree via the g3dt resolver — the job receives
only --PROJECT_ID/--ENV/--REGION from the CDK.

Environment overrides (optional): TEMPLATE_INGEST_MAX_WORKERS (default 8).
"""
import argparse
import hashlib
import io
import logging
import os
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone

import awswrangler as wr
import boto3
import openpyxl
import pandas as pd

from g3dt import resolver

logger = logging.getLogger()
logger.setLevel(logging.INFO)
handler = logging.StreamHandler(sys.stdout)
handler.setFormatter(
    logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
)
if logger.hasHandlers():
    logger.handlers.clear()
logger.addHandler(handler)

DEFAULT_MAX_WORKERS = 8
DEFAULT_PREFIX = "submissions"

#: The sheet g3mt writes its provenance and node->sheet map into.
META_SHEET = "_g3mt"
#: Row 1 headers, row 2 type/required hints, row 3 onward is data.
HEADER_ROW = 1
HINT_ROW = 2
FIRST_DATA_ROW = 3

#: Columns this job adds. Named with a leading `_src` so they cannot collide
#: with a Gen3 property (no Gen3 property starts with an underscore).
PROVENANCE_COLUMNS = [
    "_src_file",
    "_src_sheet",
    "_src_row",
    "_src_study_id",
    "_src_g3mt_version",
    "_src_schema_file",
    "_src_target_node",
    "_src_ingested_at",
    "row_hash",
]


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--PROJECT_ID", required=True)
    parser.add_argument("--ENV", required=True)
    parser.add_argument("--REGION", required=True)
    parser.add_argument(
        "--S3_PREFIX",
        default=DEFAULT_PREFIX,
        help=f"Prefix under the bronze bucket to scan (default: {DEFAULT_PREFIX}).",
    )
    parser.add_argument(
        "--STUDY",
        default=None,
        help="Ingest only this study id (default: every study found).",
    )
    parser.add_argument(
        "--DRY_RUN",
        default="false",
        help="'true' parses and reports without writing any table.",
    )
    # Glue passes --JOB_NAME and friends; ignore anything we do not declare.
    args, _unknown = parser.parse_known_args()
    return args


# ---------- workbook parsing ----------

def read_meta(workbook) -> dict:
    """Return g3mt's provenance block and node->sheet map from `_g3mt`.

    The sheet is two columns of key/value pairs. A literal ``node``/``sheet``
    pair acts as a header: everything after it is the node-to-sheet mapping,
    everything before it is scalar provenance.

    Raises:
        ValueError: if the sheet is absent — the file was not produced by
            g3mt, and guessing its layout would be worse than refusing it.
    """
    if META_SHEET not in workbook.sheetnames:
        raise ValueError(
            f"no '{META_SHEET}' sheet — this workbook was not generated by "
            "g3mt (or predates the provenance sheet), so its layout cannot be "
            "trusted. Regenerate it with `g3mt generate`."
        )

    ws = workbook[META_SHEET]
    provenance, sheets, in_map = {}, {}, False
    for row in range(1, ws.max_row + 1):
        key = ws.cell(row, 1).value
        value = ws.cell(row, 2).value
        if key is None:
            continue
        key = str(key).strip()
        if key == "node" and str(value).strip() == "sheet":
            in_map = True
            continue
        if in_map:
            sheets[key] = str(value).strip()
        else:
            provenance[key] = None if value is None else str(value).strip()

    if not sheets:
        raise ValueError(
            f"'{META_SHEET}' declares no node->sheet mapping; nothing to ingest."
        )
    return {"provenance": provenance, "sheets": sheets}


def read_node_sheet(workbook, sheet_name: str) -> pd.DataFrame:
    """Read one node sheet into a frame of strings, hint row and blanks dropped.

    Every value is kept as text on purpose. Bronze records what the researcher
    typed; coercing "2024-13-45" to a date here would either fail the whole
    ingest or silently invent a value, and neither is bronze's decision to make.
    """
    ws = workbook[sheet_name]
    headers = [
        ws.cell(HEADER_ROW, col).value for col in range(1, ws.max_column + 1)
    ]
    keep = [i for i, h in enumerate(headers) if h is not None and str(h).strip()]
    if not keep:
        return pd.DataFrame()

    names = [str(headers[i]).strip() for i in keep]
    records = []
    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        values = [ws.cell(row, i + 1).value for i in keep]
        # g3mt provisions thousands of blank rows; a row is data only if some
        # cell has content.
        if all(v is None or str(v).strip() == "" for v in values):
            continue
        record = {
            name: ("" if v is None else str(v).strip())
            for name, v in zip(names, values)
        }
        record["_src_row"] = row
        records.append(record)

    return pd.DataFrame(records, columns=names + ["_src_row"] if records else names)


def compute_row_hash(record: dict, columns: list) -> str:
    """Stable identity for a submitted row.

    Hashes the cell values **plus** the source file, sheet and row number. The
    coordinates are included so two genuinely different rows that happen to
    carry identical values (a repeated measurement, say) stay distinct, while
    re-depositing the same workbook reproduces the same hashes exactly and the
    MERGE becomes a no-op.
    """
    payload = "|".join(f"{c}={record.get(c, '')}" for c in columns)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def frame_for_node(
    workbook, sheet_name: str, node: str, meta: dict, s3_uri: str, study_id: str
) -> pd.DataFrame:
    """Parse one node sheet and stamp provenance onto every row."""
    df = read_node_sheet(workbook, sheet_name)
    if df.empty:
        return df

    provenance = meta["provenance"]
    ingested_at = datetime.now(timezone.utc).isoformat()
    data_columns = [c for c in df.columns if c != "_src_row"]

    df["_src_file"] = s3_uri
    df["_src_sheet"] = sheet_name
    df["_src_study_id"] = study_id
    df["_src_g3mt_version"] = provenance.get("g3mt_version", "")
    df["_src_schema_file"] = provenance.get("schema_file", "")
    df["_src_target_node"] = provenance.get("target_node", "")
    df["_src_ingested_at"] = ingested_at

    hash_columns = data_columns + ["_src_file", "_src_sheet", "_src_row"]
    df["row_hash"] = df.apply(
        lambda r: compute_row_hash(r.to_dict(), hash_columns), axis=1
    )
    df["_src_row"] = df["_src_row"].astype(str)
    return df


# ---------- S3 discovery ----------

def study_from_key(key: str, prefix: str) -> str:
    """First path segment under the scanned prefix is the study id."""
    rest = key[len(prefix):].lstrip("/") if key.startswith(prefix) else key
    parts = [p for p in rest.split("/") if p]
    return parts[0] if len(parts) > 1 else "unassigned"


def discover_workbooks(bucket: str, prefix: str, session, study=None) -> list:
    """List candidate .xlsx objects, optionally narrowed to one study."""
    uris = wr.s3.list_objects(f"s3://{bucket}/{prefix.rstrip('/')}/", boto3_session=session)
    found = []
    for uri in uris:
        key = uri.split(f"s3://{bucket}/", 1)[-1]
        name = key.rsplit("/", 1)[-1]
        # Excel writes ~$lockfiles next to open workbooks; never ingest them.
        if not name.lower().endswith(".xlsx") or name.startswith("~$"):
            continue
        found_study = study_from_key(key, prefix.rstrip("/"))
        if study and found_study != study:
            continue
        found.append((uri, found_study))
    return found


def load_workbook_from_s3(uri: str, session):
    bucket, key = uri.replace("s3://", "", 1).split("/", 1)
    body = session.client("s3").get_object(Bucket=bucket, Key=key)["Body"].read()
    # read_only=False: the provenance sheet is small and read_only mode reports
    # unreliable max_row on some g3mt-written sheets.
    return openpyxl.load_workbook(io.BytesIO(body), data_only=True)


# ---------- main ----------

def ingest_workbook(uri: str, study_id: str, session) -> list:
    """Parse one workbook into (table_name, frame) pairs. Raises on bad input."""
    workbook = load_workbook_from_s3(uri, session)
    meta = read_meta(workbook)
    results = []
    for node, sheet_name in meta["sheets"].items():
        if sheet_name not in workbook.sheetnames:
            logger.warning(
                "%s: '_g3mt' maps node '%s' to missing sheet '%s'; skipping.",
                uri, node, sheet_name,
            )
            continue
        df = frame_for_node(workbook, sheet_name, node, meta, uri, study_id)
        if df.empty:
            logger.info("%s: node '%s' has no filled rows.", uri, node)
            continue
        results.append((f"bronze_{study_id}_{node}", df))
    return results


def main():
    args = parse_args()
    dry_run = str(args.DRY_RUN).strip().lower() == "true"
    session = boto3.Session(region_name=args.REGION)

    rc = resolver.ResolvedConfig(args.PROJECT_ID, args.ENV, args.REGION)
    bronze_bucket = rc.get("buckets/bronze")
    bronze_db = rc.get("glue/db/bronze")
    athena_output = rc.get("athena/output")
    workgroup = rc.get("athena/workgroup")

    prefix = args.S3_PREFIX.strip("/")
    logger.info(
        "Scanning s3://%s/%s/ for g3mt workbooks (study filter: %s)",
        bronze_bucket, prefix, args.STUDY or "ALL",
    )

    workbooks = discover_workbooks(bronze_bucket, prefix, session, args.STUDY)
    if not workbooks:
        logger.warning("No .xlsx workbooks found. Nothing to do.")
        return
    logger.info("Found %d workbook(s).", len(workbooks))

    max_workers = int(os.environ.get("TEMPLATE_INGEST_MAX_WORKERS", DEFAULT_MAX_WORKERS))
    parsed, failures = {}, []

    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        futures = {
            pool.submit(ingest_workbook, uri, study, boto3.Session(region_name=args.REGION)): uri
            for uri, study in workbooks
        }
        for future in as_completed(futures):
            uri = futures[future]
            try:
                for table, df in future.result():
                    parsed.setdefault(table, []).append(df)
                    logger.info("%s -> %s (%d row(s))", uri, table, len(df))
            except Exception as exc:  # noqa: BLE001 — reported per workbook below
                logger.error("FAILED %s: %s", uri, exc)
                failures.append((uri, str(exc)))

    if not parsed:
        if failures:
            raise RuntimeError(
                f"No workbook parsed successfully; {len(failures)} failed."
            )
        logger.warning("No filled rows in any workbook.")
        return

    for table, frames in sorted(parsed.items()):
        df = pd.concat(frames, ignore_index=True)
        # Two workbooks in one run can carry the same row; keep one.
        df = df.drop_duplicates(subset=["row_hash"])
        if dry_run:
            logger.info("[dry run] %s.%s would receive %d row(s), columns: %s",
                        bronze_db, table, len(df), list(df.columns))
            continue
        logger.info("Writing %d row(s) to %s.%s", len(df), bronze_db, table)
        wr.athena.to_iceberg(
            df=df,
            database=bronze_db,
            table=table,
            table_location=f"s3://{bronze_bucket}/{table}/",
            temp_path=f"{athena_output.rstrip('/')}/tmp_{table}/",
            workgroup=workgroup,
            # MERGE on row_hash: re-depositing an unchanged workbook rewrites
            # the same rows rather than appending a second copy.
            merge_cols=["row_hash"],
            schema_evolution=True,
            boto3_session=session,
        )

    if failures:
        raise RuntimeError(
            "Ingest finished with %d failed workbook(s): %s"
            % (len(failures), "; ".join(f"{u} ({e})" for u, e in failures))
        )
    logger.info("Ingest complete: %d bronze table(s) written.", len(parsed))


if __name__ == "__main__":
    main()
